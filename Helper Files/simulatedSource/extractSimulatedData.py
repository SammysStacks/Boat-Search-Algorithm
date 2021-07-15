#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jun 29 17:00:25 2021

@author: samuelsolomon
"""


# Basic Modules
import os
import sys
import numpy as np
# Read/Write to Excel
import csv
import pyexcel
import openpyxl as xl
# Plotting
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import axes3d
# Interpolate
from scipy import interpolate


class dataProcessing:        
        
    def convertToXLSX(self, excelFile):
        """
        Converts .xls Files to .xlsx Files That OpenPyxl Can Read
        If the File is Already a .xlsx Files, Do Nothing
        If the File is Neither a .xls Nor .xlsx, it Exits the Program
        """
        # Check That the Current Extension is .xls or .xlsx
        _, extension = os.path.splitext(excelFile)
        # If the Extension is .xlsx, the File is Ready; Do Nothing
        if extension == '.xlsx':
            return excelFile
        # If the Extension is Not .xls/.xlsx, Then the Data is in the Wrong Format; Exit Program
        if extension not in ['.xls', '.xlsx']:
            print("Cannot Convert File to .xlsx")
            sys.exit()
        
        # Create Output File Directory to Save Data ONLY If None Exists
        newExcelFolder = os.path.dirname(excelFile) + "/Excel Files/"
        os.makedirs(newExcelFolder, exist_ok = True)
        
        # Convert '.xls' to '.xlsx'
        filename = os.path.basename(excelFile)
        newExcelFile = newExcelFolder + filename + "x"
        pyexcel.save_as(file_name = excelFile, dest_file_name = newExcelFile, logfile=open(os.devnull, 'w'))
        
        # Save New Excel name
        return newExcelFile
    
    def txt2csv(self, txtFile, csvFile, csvDelimiter = ",", overwriteCSV = False):
        # Check to see if csv conversion alreayd happened
        if not os.path.isfile(csvFile) or overwriteCSV:
            with open(txtFile, "r") as inputData:
                in_reader = csv.reader(inputData, delimiter = csvDelimiter)
                with open(csvFile, 'w', newline='') as out_csv:
                    out_writer = csv.writer(out_csv)
                    for row in in_reader:
                        out_writer.writerow(row)
    
    def convertToExcel(self, inputFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = 0):
        # If the File is Not Already Converted: Convert the CSV to XLSX
        if not os.path.isfile(excelFile) or overwriteXL:
            # Make Excel WorkBook
            xlWorkbook = xl.Workbook()
            xlWorksheet = xlWorkbook.active
            # Write the Data from the CSV File to the Excel WorkBook
            with open(inputFile, "r") as inputData:
                inReader = csv.reader(inputData, delimiter = excelDelimiter)
                with open(excelFile, 'w+', newline=''):
                    for row in inReader:
                        xlWorksheet.append(row)
            # Save as New Excel File
            xlWorkbook.save(excelFile)
        # Else Load the GSR Data from the Excel File
        else:
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        
        # Return Excel Sheet
        return xlWorkbook, xlWorksheet
    

class processData(dataProcessing):
    
    def extractCosmolData(self, xlWorksheet, yVal = 25, zCol = 7):
        
        # -------------------------------------------------------------------#
        # ----------------------- Extract Run Info --------------------------#
        
        x = []; z = []; concentrations = []
        # Loop Through the Info Section and Extract the Needed Run Info from Excel
        rowGenerator = xlWorksheet.rows
        for cell in rowGenerator:
            
            if cell[1].value == yVal:
                x.append(cell[0].value)
                z.append(cell[2].value)
                concentrations.append(cell[zCol].value)
        
        return x, z, concentrations
    
    def getData(self, oldFile, testSheetNum = 0):
        """
        --------------------------------------------------------------------------
        Input Variable Definitions:
            excelFile: The Path to the Excel File Containing the Data
            testSheetNum: An Integer Representing the Excel Worksheet (0-indexed) Order.
        --------------------------------------------------------------------------
        """
        # Check if File Exists
        if not os.path.exists(oldFile):
            print("The following Input File Does Not Exist:", oldFile)
            sys.exit()
            
        # Convert TXT and CSV Files to XLSX
        if oldFile.endswith((".txt", ".csv", ".numbers")):
            # Extract Filename Information
            oldFileExtension = os.path.basename(oldFile)
            filename = os.path.splitext(oldFileExtension)[0]
            newFilePath = os.path.dirname(oldFile) + "/Excel Files/"
            # Make Output Folder Directory if Not Already Created
            os.makedirs(newFilePath, exist_ok = True)

            # Convert CSV or TXT to XLSX
            excelFile = newFilePath + filename + ".xlsx"
            xlWorkbook, xlWorksheet = self.convertToExcel(oldFile, excelFile, excelDelimiter = ",", overwriteXL = False, testSheetNum = testSheetNum)
        # If the File is Already an Excel File, Just Load the File
        elif oldFile.endswith(".xlsx"):
            excelFile = oldFile
            # Load the GSR Data from the Excel File
            xlWorkbook = xl.load_workbook(excelFile, data_only=True, read_only=True)
            xlWorksheet = xlWorkbook.worksheets[testSheetNum]
        else:
            print("The Following File is Neither CSV, TXT, Nor XLSX:", oldFile)
            sys.exit()
        print("Extracting Data from the Excel File:", excelFile)
        
        # Extract Time and Current Data from the File
        xPoints, zPoints, concentrations = self.extractCosmolData(xlWorksheet, yVal = 25)
        
        xlWorkbook.close()
        # Finished Data Collection: Close Workbook and Return Data to User
        print("Done Collecting Data");
        return np.array(xPoints), np.array(zPoints), np.array(concentrations)


if __name__ == "__main__":
    
    cosmolFile = './Input Data/diffusion4.xlsx'
    x, y, z = processData().getData(cosmolFile)
    # Rescale Data
    if True:
        x -= min(x)
        y -= min(y)
        # Reduce X,Y to Gameboard Positions
        x = x*20/max(x)
        y = y*20/max(y)
    
    # Get Sampled Data
    xSamples = np.arange(min(x), max(x), 0.2)
    ySamples = np.arange(min(y), max(y), 0.2)
    xx,yy=np.meshgrid(xSamples, ySamples)
    
    # Interpolate Sampled Data
    zSamples = interpolate.griddata((x, y), z, (xx,yy), method='cubic')
    
    # Plot Data
    fig = plt.figure()
    ax = axes3d.Axes3D(fig)
    ax.scatter(xx.ravel(), yy.ravel(), zSamples.ravel(), c=zSamples.ravel())
    plt.show()
    
    # Plot Model
    fig = plt.figure()
    ax = axes3d.Axes3D(fig)
    ax.scatter(x, y, z, c=z)
    ax.scatter(x[z<0],y[z<0],z[z<0], c = 'black')
    plt.show()
    
    
    